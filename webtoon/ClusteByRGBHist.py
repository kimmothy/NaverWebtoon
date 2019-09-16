import numpy as np
import cv2
import random
import openpyxl as op

class toon:
    def __init__(self,num,hist):
        self.num=num
        self.hist=hist

# getCenter, getClosest, getDistance 세 함수는 k-means를 사용해 웹툰을 구분하기 위해 쓰이는 함수
def getCenter(cluster):
    n = len(cluster)
    base = np.zeros(384)
    for i in cluster:
        base = base + i.hist
    return base/n

def getClosest(t,centers):
    k = len(centers)
    closestDistance = getDistance(t.hist,centers[0])
    closestCenterIndex = 0
    for i in range(1,k):
        newDistance = getDistance(t.hist,centers[i]) 
        if newDistance < closestDistance:
            closestDistance=newDistance
            closestCenterIndex = i
    return closestCenterIndex


def getDistance(h1, h2):
    tmpHist = (h1-h2)*(h1-h2)
    tmp = 0
    for i in range(384):
        tmp += tmpHist[i]
    return np.sqrt(tmp)

# openCV와 numPy를 사용해 색 사용량을 히스토그램으로 변환하는 함수(R, G, B 따로)
def getRGBArray(imgsrc):
    img_b, img_g, img_r = cv2.split(imgsrc)
    hist_b = np.histogram(img_b, 128, [0,256])[0]
    hist_g = np.histogram(img_g, 128, [0,256])[0]
    hist_r = np.histogram(img_r, 128, [0,256])[0]
    hist_bgr = np.append(hist_b,[hist_g,hist_r])
    return hist_bgr

# NaWebCrolling.py를 사용해 수집해놓은 웹툰 썸네일 이미지 로드
# 중간 if 문은 cv2.imrad가 gif 파일을 읽어오지 못하기때문에, 통과하도록 한 것
# gif파일을 jpg로 인코딩후 실생하면 모든 이미지 분석 가능
toons = []
for i in range(2,837):
    img = cv2.imread('C:\python_data\\webtoon\\imgs\\'+str(i)+'.jpg', cv2.IMREAD_COLOR)
    if img is None:
        continue
    hist = getRGBArray(img)
    t = toon(i,hist)
    toons.append(t)

# 이 이하는 k-means기법을 사용해 웹툰 분류, k 변수는 분류할 클러스터의 갯수
k = 5
centerIndex = random.sample(range(0,len(toons)),k)
centers = []
for i in centerIndex:
    centers.append(toons[i].hist)


clusters = []
while True:
    newClusters = [[] for x in range(k)]
    for t in toons:
        closestCenterIndex = getClosest(t,centers)
        newClusters[closestCenterIndex].append(t)
    for j in range(k):
        centers[j]=getCenter(newClusters[j])
    if newClusters == clusters:
        break
    else:
        clusters = newClusters
        continue

#분석내용 출력

wb = op.load_workbook('C:\python_data\\webtoon\\webtooninfo.xlsx')
ws = wb.get_sheet_by_name('sheet1')
for i in range(len(clusters)):
    print("\n\n 여기서부터는 클러스터"+str(i+1)+"에 속한 만화")
    clusterNames = ""
    for t in clusters[i]:
        nameCell = 'A' + str(t.num)
        toonName = ws[nameCell].value
        clusterNames = clusterNames + toonName + " "
    print(clusterNames)

for i in range(len(clusters)):
    for t in clusters[i]:
        index = 'F' + str(t.num)
        ws[index]='클러스터' + str(i+1)
wb.save('C:\python_data\\webtoon\\webtooninfo.xlsx')
wb.close()
