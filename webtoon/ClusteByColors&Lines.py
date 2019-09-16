import cv2
import numpy as np
import random
import openpyxl as op

class toon:
    def __init__(self,num,lines,colors):
        self.num=num
        self.lines=lines
        self.colors=colors*3

# getCenter, getClosest, getDistance 세 함수는 k-means를 사용해 웹툰을 구분하기 위해 쓰이는 함수
def getCenter(cluster):
    n = len(cluster)
    baseColor = 0
    baseLine = 0
    for i in cluster:
        baseLine = baseLine + i.lines
        baseColor = baseColor + i.colors
    return (baseLine/n, baseColor/n)

def getClosest(t,centers):
    k = len(centers)
    ttuple = (t.lines,t.colors)
    closestDistance = getDistance(ttuple,centers[0])
    closestCenterIndex = 0
    for i in range(1,k):
        newDistance = getDistance(ttuple,centers[i]) 
        if newDistance < closestDistance:
            closestDistance=newDistance
            closestCenterIndex = i
    return closestCenterIndex

def getDistance(t1, t2):
    tmpLine = (t1[0]-t2[0])*(t1[0]-t2[0])
    tmpColor = (t1[1]-t2[1])*(t1[1]-t2[1])
    tmp = tmpLine + tmpColor
    return np.sqrt(tmp)

# getLines는 그림에 얼마나 많은 색을 사용했는지 파악하는 함수
# cv2.Canny 함수는 그림에서 선으로 인식되는 부분을 검은색(255), 나머지를 흰색(0)으로 바꿔준다.
# 처리된 이미지 히스토그램에서 검은 색(255) 픽셀의 갯수를 반환한다.
def getLines(img):
    edge = cv2.Canny(img, 100, 200)
    return np.histogram(edge, 256, [0,256])[0][255]

# getColors는 그림에서 얼마나 다양한 색을 사용했는지 파악하는 함수
# 각 색의 강도별로 쓰인 픽셀수를 파악, 10픽셀 이상 쓰인 색의 갯수를 반환한다.
def getColors(imgsrc):
    img_b, img_g, img_r = cv2.split(imgsrc)
    hist_b = np.histogram(img_b, 256, [0,256])[0]
    hist_g = np.histogram(img_g, 256, [0,256])[0]
    hist_r = np.histogram(img_r, 256, [0,256])[0]
    hist_bgr = np.append(hist_b,[hist_g,hist_r])
    
    count = 0    
    for i in hist_bgr:
        if i > 10:
            count += 1
    return count

toons =[]

# NaWebCrolling.py를 사용해 수집해놓은 웹툰 썸네일 이미지 로드
# 중간 if 문은 cv2.imrad가 gif 파일을 읽어오지 못하기때문에, 통과하도록 한 것
# gif파일을 jpg로 인코딩후 실생하면 모든 이미지 분석 가능
for i in range(2,837):
    img = cv2.imread('C:\python_data\\webtoon\\imgs\\'+str(i)+'.jpg', cv2.IMREAD_COLOR)
    if img is None:
        continue
    colors = getColors(img)
    lines = getLines(img)
    t = toon(i,lines,colors)
    toons.append(t)

allLines =[]
allColors =[]

for t in toons:
    allLines.append(t.lines)
    allColors.append(t.colors)
allLines.sort()
allColors.sort()

toonNum = len(toons)
AverageLine = 0
for i in allLines:
    AverageLine += i
AverageLine = AverageLine//toonNum

AverageColor = 0
for i in allColors:
    AverageColor += i
AverageColor = AverageColor//toonNum
print("선 통계")
print("Max:",str(allLines[toonNum-1]),"Min:",str(allLines[0]),"Average:",AverageLine, "Quarter:",str(allLines[toonNum//4]),"mid:",str(allLines[toonNum//2]),"3Quarter:",str(allLines[toonNum//4*3]),sep=" ")
print("색 통계")
print("Max:",str(allColors[toonNum-1]),"Min:",str(allColors[0]),"Average:",AverageColor, "Quarter:",str(allColors[toonNum//4]),"mid:",str(allColors[toonNum//2]),"3Quarter:",str(allColors[toonNum//4*3]),sep=" ")

# 이 이하는 k-means기법을 사용해 웹툰 분류, k 변수는 분류할 클러스터의 갯수
k = 5
centerIndex = random.sample(range(0,835),k)
centers = []
for i in centerIndex:
    cLine = toons[i].lines
    cColor = toons[i].colors
    center = (cLine, cColor)
    centers.append(center)

clusters = []
while True:
    newClusters = [[] for x in range(k)]
    for t in toons:
        closestCenterIndex = getClosest(t,centers)
        newClusters[closestCenterIndex].append(t)
    for j in range(k):
        centers[j]=getCenter(newClusters[j])
    if newClusters == clusters:
        break
    else:
        clusters = newClusters
        continue
print(centers)

#분석내용 출력

wb = op.load_workbook('C:\python_data\\webtoon\\webtooninfo.xlsx')
ws = wb.get_sheet_by_name('sheet1')
for i in range(len(clusters)):
    print("\n\n 여기서부터는 클러스터"+str(i+1)+"에 속한 만화")
    clusterNames = ""
    for t in clusters[i]:
        nameCell = 'A' + str(t.num)
        toonName = ws[nameCell].value
        clusterNames = clusterNames + toonName + " "
    print(clusterNames)


for i in range(len(clusters)):
    for t in clusters[i]:
        index = 'E' + str(t.num)
        ws[index]='클러스터' + str(i+1)
wb.save('C:\python_data\\webtoon\\webtooninfo.xlsx')
wb.close()
