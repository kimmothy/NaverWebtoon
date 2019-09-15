from selenium import webdriver
import openpyxl as op
from openpyxl import Workbook
import urllib.request
import re


# 장르별 분류로 들어가 각 웹툰의 장르 가져오는 함수
# 첫번째 반복문은 장르별로 제시된 웹툰들의 이름을 가져온다
# 두번째 반복문은 각 웹툰의 이름을 검색해 자신의 장르를 저장한다.
def getGenres():
    genres = ['epsode', 'omnibus', 'story', 'daily', 'comic', 'fantasy', 'action', 'drama', 'pure', 'sensibility', 'thrill', 'historical', 'sports']
    baseurl = 'https://comic.naver.com/webtoon/genre.nhn?genre='
    totalIndex = {}
    for i in genres:
        browser.get(baseurl + i)
        toons = browser.find_elements_by_css_selector('ul.img_list > li')
        genreIndex = []
        for j in toons:
            toon_name = j.find_element_by_css_selector('dl > dt > a').get_attribute('title')
            filter = re.compile('[^ 0123456789A-Za-zㄱ-ㅣ가-힣]+')
            toon_name = filter.sub('', toon_name)
            genreIndex.append(toon_name)
        totalIndex[i] = genreIndex
    wb = op.load_workbook('C:\python_data\\webtoon\\webtooninfo.xlsx')
    ws = wb.get_sheet_by_name('sheet1')
    ws['D1'] = '장르'
    numOfToon = ws.max_row
    for i in range(2, numOfToon+1):
        genresOfThis = ''
        nameCell = 'A' + str(i)
        genreCell = 'D' + str(i)
        toonName = ws[nameCell].value
        for j in genres:
            if toonName in totalIndex[j]:
                genresOfThis = genresOfThis + j +', '
        genresOfThis = genresOfThis[:-2]
        ws[genreCell] = genresOfThis
    wb.save('C:\python_data\\webtoon\\webtooninfo.xlsx')
    wb.close()


def WriteToExcel(name, author,rate):
    wb = op.load_workbook('C:\python_data\\webtoon\\webtooninfo.xlsx')
    ws = wb.get_sheet_by_name('sheet1')
    last_index = ws.max_row
    A = 'A' + str(last_index + 1)
    B = 'B' + str(last_index + 1)
    C = 'C' + str(last_index + 1)
    ws[A] = name
    ws[B] = author
    ws[C] = rate

    wb.save('C:\python_data\\webtoon\\webtooninfo.xlsx')
    wb.close()


def makeExcel():
    wb = Workbook()
    ws = wb.active
    ws.title = 'sheet1'
    ws['A1'] = '이름'
    ws['B1'] = '작가'
    ws['C1'] = '별점'
    wb.save('C:\python_data\\webtoon\\webtooninfo.xlsx')
    wb.close()

# selenium의 태그 객체(각 웹툰 정보 부분)을 받아, 썸네일이미지를 다운로드하고, 이름, 작가 별점을 엑셀에 저장
def getInfos(toon):
    global num
    toon_name = toon.find_element_by_css_selector('dl > dt > a').get_attribute('title')
    filter = re.compile('[^ 0123456789A-Za-zㄱ-ㅣ가-힣]+')
    toon_name = filter.sub('', toon_name)
    if toon_name in index:
        return
    else:
        imgsrc = toon.find_element_by_css_selector('img').get_attribute('src')
        if imgsrc[-3:] == 'gif':
            urllib.request.urlretrieve(imgsrc, 'C:\python_data\webtoon\imgs\\'+str(num)+'.gif')
        else:
            urllib.request.urlretrieve(imgsrc, 'C:\python_data\webtoon\imgs\\'+str(num)+'.jpg')
        author = toon.find_element_by_css_selector('dd.desc > a').text
        star = toon.find_element_by_css_selector('div.rating_type > strong').text
        WriteToExcel(toon_name, author, star)
        index.append(toon_name)
        num += 1

browser = webdriver.Chrome('C:\python_data\chromedriver_win32\chromedriver.exe')

baseurl = 'https://comic.naver.com/webtoon/weekdayList.nhn?week='
browser.implicitly_wait(3)
weekdays = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun']

index=[]

makeExcel()
num = 2

for i in weekdays:
    browser.get(baseurl + i)
    toons = browser.find_elements_by_css_selector('ul.img_list > li')
    for j in toons:
        getInfos(j)

browser.get('https://comic.naver.com/webtoon/finish.nhn')
finished = browser.find_elements_by_css_selector('ul.img_list > li')
print(len(finished))

for j in finished:
    getInfos(j)

getGenres()